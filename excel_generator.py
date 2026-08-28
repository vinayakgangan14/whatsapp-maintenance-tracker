import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import datetime
from config import EXPORTS_DIR
import database

def generate_excel_report():
    """
    Generates a professionally styled Excel workbook containing
    Breakdown Logs, Preventive Maintenance Logs, and MTTR Statistics.
    Returns file_path.
    """
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # TAB 1: BREAKDOWNS LOG
    # ----------------------------------------------------
    ws_bd = wb.active
    ws_bd.title = "Breakdown Tracking"
    ws_bd.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws_bd.merge_cells("A1:J1")
    title_cell = ws_bd["A1"]
    title_cell.value = "PURE BOT FOR MAINTENANCE AND REPAIR — BREAKDOWN TRACKING REPORT"
    title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_bd.row_dimensions[1].height = 40

    # Timestamp & Metrics Row
    stats = database.get_statistics()
    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    ws_bd.merge_cells("A2:J2")
    info_cell = ws_bd["A2"]
    info_cell.value = f"Generated: {now_str} | Open: {stats['open_breakdowns']} | Resolved: {stats['resolved_breakdowns']} | Total Downtime: {stats['total_downtime_hours']} hrs | MTTR: {stats['mttr_minutes']} mins"
    info_cell.font = Font(name="Calibri", size=10, italic=True, color="595959")
    info_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_bd.row_dimensions[2].height = 20

    # Header Row
    headers = [
        "Ticket Number", "Department", "Equipment ID", "Issue Description", 
        "Status", "Start Time", "End Time", "Duration (Mins)", "Resolution Notes", "Technician"
    ]
    
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws_bd.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws_bd.row_dimensions[4].height = 28

    # Populate Data
    breakdowns = database.get_all_breakdowns(limit=500)
    
    open_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Soft Red
    resolved_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft Green
    open_font = Font(name="Calibri", size=10, bold=True, color="C00000")
    resolved_font = Font(name="Calibri", size=10, bold=True, color="375623")
    regular_font = Font(name="Calibri", size=10, color="000000")

    for row_idx, item in enumerate(breakdowns, start=5):
        ws_bd.cell(row=row_idx, column=1, value=item['ticket_number']).alignment = Alignment(horizontal="center")
        ws_bd.cell(row=row_idx, column=2, value=item['department']).alignment = Alignment(horizontal="center")
        ws_bd.cell(row=row_idx, column=3, value=item['equipment_id']).alignment = Alignment(horizontal="left")
        ws_bd.cell(row=row_idx, column=4, value=item['issue_description']).alignment = Alignment(horizontal="left")
        
        status_cell = ws_bd.cell(row=row_idx, column=5, value=item['status'])
        status_cell.alignment = Alignment(horizontal="center")
        if item['status'] == 'OPEN':
            status_cell.fill = open_fill
            status_cell.font = open_font
        else:
            status_cell.fill = resolved_fill
            status_cell.font = resolved_font
            
        ws_bd.cell(row=row_idx, column=6, value=item['start_time'][:16].replace('T', ' ')).alignment = Alignment(horizontal="center")
        ws_bd.cell(row=row_idx, column=7, value=(item['end_time'] or '')[:16].replace('T', ' ')).alignment = Alignment(horizontal="center")
        
        dur_cell = ws_bd.cell(row=row_idx, column=8, value=item['duration_minutes'])
        dur_cell.alignment = Alignment(horizontal="right")
        
        ws_bd.cell(row=row_idx, column=9, value=item['resolution_notes'] or '-').alignment = Alignment(horizontal="left")
        ws_bd.cell(row=row_idx, column=10, value=item['technician'] or item['sender_name'] or '-').alignment = Alignment(horizontal="left")

        # Apply borders and font
        for c in range(1, 11):
            cell = ws_bd.cell(row=row_idx, column=c)
            cell.border = thin_border
            if c != 5: # Keep custom status font
                cell.font = regular_font

        ws_bd.row_dimensions[row_idx].height = 22

    # Auto-adjust column widths
    for col in ws_bd.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_bd.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # ----------------------------------------------------
    # TAB 2: PREVENTIVE MAINTENANCE LOG
    # ----------------------------------------------------
    ws_pm = wb.create_sheet(title="Preventive Maintenance")
    ws_pm.views.sheetView[0].showGridLines = True
    
    ws_pm.merge_cells("A1:F1")
    pm_title = ws_pm["A1"]
    pm_title.value = "PREVENTIVE MAINTENANCE (PM) ACTIVITY LOG"
    pm_title.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    pm_title.fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
    pm_title.alignment = Alignment(horizontal="center", vertical="center")
    ws_pm.row_dimensions[1].height = 35

    pm_headers = ["Ticket Number", "Department", "Equipment ID", "Activity Description", "Technician", "Performed At"]
    for col_num, header in enumerate(pm_headers, 1):
        cell = ws_pm.cell(row=3, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws_pm.row_dimensions[3].height = 26

    pm_logs = database.get_all_maintenance(limit=500)
    for row_idx, item in enumerate(pm_logs, start=4):
        ws_pm.cell(row=row_idx, column=1, value=item['ticket_number']).alignment = Alignment(horizontal="center")
        ws_pm.cell(row=row_idx, column=2, value=item['department']).alignment = Alignment(horizontal="center")
        ws_pm.cell(row=row_idx, column=3, value=item['equipment_id']).alignment = Alignment(horizontal="left")
        ws_pm.cell(row=row_idx, column=4, value=item['activity_description']).alignment = Alignment(horizontal="left")
        ws_pm.cell(row=row_idx, column=5, value=item['technician']).alignment = Alignment(horizontal="left")
        ws_pm.cell(row=row_idx, column=6, value=item['performed_at'][:16].replace('T', ' ')).alignment = Alignment(horizontal="center")

        for c in range(1, 7):
            cell = ws_pm.cell(row=row_idx, column=c)
            cell.border = thin_border
            cell.font = regular_font
        ws_pm.row_dimensions[row_idx].height = 20

    for col in ws_pm.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_pm.column_dimensions[col_letter].width = max(max_len + 4, 14)

    # Save to file
    filename = f"Maintenance_Report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = EXPORTS_DIR / filename
    wb.save(filepath)
    return str(filepath), filename
